#excel_tbs.py by Author: Hyslan Silva Cruz
'''Módulo de acessar arquivos .xlsx'''
#Bibliotecas
from openpyxl import load_workbook # Carregar função load

#Área do Excel
def load_worksheets():
    '''Função para carregar os arquivos dos Excel e futuramente do SQL Server.'''    
    lista = load_workbook('lista.xlsx') # Carregando arquivo para valorar
    materiais = load_workbook('MateriaisContratada.xlsx') # Tabela de materiais da Contratada
    plan_tse = load_workbook('TSE.xlsx') # Tabela de TSE
    plan_precos = load_workbook('ItensPreco.xlsx') # Tabela Itens de preço
    plan_st_ordem = load_workbook('StatusOrdem.xlsx') # Tabela Status da Ordem
    planilha = lista.active # Planilha de Ordem
    contratada = materiais.active  # Materiais Contratada
    st_sistema = plan_st_ordem["StatusSistema"] # Status Sistema
    st_usuario = plan_st_ordem["StatusUsuario"] # Status Usuario
    unitario = plan_tse["unitario"]
    rem_base = plan_tse["rb"]
    naoexecutado = plan_tse["NEXEC"]
    invest = plan_tse["invest"]
    n_motivo3 = plan_tse["n3"]
    n10 = plan_tse["n10"]
    reposicao = plan_tse["reposicao"]
    retrabalho = plan_tse["retrabalho"]
    asfalto = plan_tse["asfalto"]
    coluna_contratada = 'A'
    coluna_tse = 'A'
    coluna_status = 'A'

    tb_contratada = [cell.value for cell in contratada[coluna_contratada]]

    plan_st_ordem.active = st_sistema
    tb_st_sistema = [cell.value for cell in st_sistema[coluna_status]]

    plan_st_ordem.active = st_usuario
    tb_st_usuario = [cell.value for cell in st_sistema[coluna_status]]

    plan_tse.active = unitario
    tb_tse_un = [cell.value for cell in unitario[coluna_tse]]

    plan_tse.active = rem_base
    tb_tse_rem_base = [cell.value for cell in rem_base[coluna_tse]]

    plan_tse.active = naoexecutado
    tb_tse_nexec = [cell.value for cell in naoexecutado[coluna_tse]]

    plan_tse.active = invest
    tb_tse_invest = [cell.value for cell in invest[coluna_tse]]

    plan_tse.active = n_motivo3
    tb_tse_pertence_ao_servico_principal = [cell.value for cell in n_motivo3[coluna_tse]]

    plan_tse.active = n10
    tb_tse_servico_nao_existe_no_contrato = [cell.value for cell in n10[coluna_tse]]

    plan_tse.active = reposicao
    tb_tse_reposicao = [cell.value for cell in reposicao[coluna_tse]]

    plan_tse.active = retrabalho
    tb_tse_retrabalho = [cell.value for cell in retrabalho[coluna_tse]]

    plan_tse.active = asfalto
    tb_tse_asfalto = [cell.value for cell in asfalto[coluna_tse]]

    return (
        lista,
        materiais,
        plan_tse,
        plan_precos,
        planilha,
        contratada,
        unitario,
        rem_base,
        naoexecutado,
        invest,
        tb_contratada,
        tb_tse_un,
        tb_tse_rem_base,
        tb_tse_nexec,
        tb_tse_invest,
        tb_st_sistema,
        tb_st_usuario,
        tb_tse_pertence_ao_servico_principal,
        tb_tse_servico_nao_existe_no_contrato,
        tb_tse_reposicao,
        tb_tse_retrabalho,
        tb_tse_asfalto,
            )
    